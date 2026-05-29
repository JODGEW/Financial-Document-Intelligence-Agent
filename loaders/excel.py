"""Excel format handler.

Uses openpyxl directly so we control sheet→Document mapping and chunk size.
Each sheet becomes its own ``Document`` prefixed with a human-readable
description block (filename, sheet name, row count, column names) so retrieval
has semantic anchors for tabular content. Oversized sheets are split into
``EXCEL_MAX_ROWS_PER_CHUNK`` chunks at split time, with the description and
header repeated on every chunk. ``format_family="tabular"`` makes the handler
eligible for ingest-time PII redaction.
"""

from __future__ import annotations

from pathlib import Path

from langchain_core.documents import Document

from .common import humanize_stem
from .registry import FormatHandler, register
from .tables import format_table_rows


EXCEL_MAX_ROWS_PER_CHUNK = 50


def _description(path: str, sheet_name: str, columns: list, row_count: int) -> str:
    human = humanize_stem(path)
    cleaned_cols = [str(c).strip() for c in columns if c is not None and str(c).strip()]
    cols = ", ".join(cleaned_cols) if cleaned_cols else "(no columns detected)"
    return (
        f"Document: {human}\n"
        f"File: {Path(path).name}\n"
        f"Sheet: {sheet_name}\n"
        f"Rows: {row_count}\n"
        f"Columns: {cols}"
    )


def _load(path: str) -> list[Document]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True, read_only=True)
    docs: list[Document] = []
    try:
        for sheet in workbook.worksheets:
            rows: list[list[object]] = []
            for row in sheet.iter_rows(values_only=True):
                rows.append(list(row))
            text = format_table_rows(rows)
            if not text:
                continue
            columns = rows[0] if rows else []
            body_count = max(0, len(rows) - 1)
            description = _description(path, sheet.title, columns, body_count)
            docs.append(
                Document(
                    page_content=f"{description}\n\n{text}",
                    metadata={
                        "source": path,
                        "section_title": f"Sheet: {sheet.title}",
                        "sheet_name": sheet.title,
                        "sheet_row_count": len(rows),
                    },
                )
            )
    finally:
        workbook.close()
    return docs


def _slice_lines(lines: list[str]) -> tuple[list[str], list[str], list[str]]:
    """Return (description_prefix, table_header, body_rows)."""
    table_start = None
    for i, line in enumerate(lines):
        if line.startswith("|"):
            table_start = i
            break

    if table_start is None:
        return lines, [], []

    prefix = lines[:table_start]
    rest = lines[table_start:]
    if len(rest) >= 2 and rest[1].startswith("|") and "---" in rest[1]:
        return prefix, rest[:2], rest[2:]
    return prefix, [], rest


def _split(docs: list[Document]) -> list[Document]:
    """Cap each sheet to ``EXCEL_MAX_ROWS_PER_CHUNK`` rows per chunk.

    Description prefix and table header are repeated on every chunk so column
    labels and the document context survive across chunks.
    """
    chunks: list[Document] = []
    for doc in docs:
        lines = doc.page_content.split("\n")
        prefix, header_block, body_lines = _slice_lines(lines)
        while prefix and not prefix[-1].strip():
            prefix.pop()

        if len(body_lines) <= EXCEL_MAX_ROWS_PER_CHUNK:
            chunks.append(doc)
            continue

        total_chunks = (len(body_lines) + EXCEL_MAX_ROWS_PER_CHUNK - 1) // EXCEL_MAX_ROWS_PER_CHUNK
        if total_chunks > 1:
            print(
                f"  INFO: sheet {doc.metadata.get('sheet_name')!r} split into "
                f"{total_chunks} chunks (rows={len(body_lines)}, cap={EXCEL_MAX_ROWS_PER_CHUNK})"
            )
        for chunk_index in range(total_chunks):
            start = chunk_index * EXCEL_MAX_ROWS_PER_CHUNK
            end = start + EXCEL_MAX_ROWS_PER_CHUNK
            assembled: list[str] = []
            if prefix:
                assembled.extend(prefix)
                assembled.append("")
            assembled.extend(header_block)
            assembled.extend(body_lines[start:end])
            metadata = {**doc.metadata, "sheet_chunk_index": chunk_index}
            chunks.append(Document(page_content="\n".join(assembled), metadata=metadata))
    return chunks


register(
    FormatHandler(
        extensions=(".xlsx", ".xls"),
        loader=_load,
        splitter=_split,
        format_family="tabular",
    )
)
